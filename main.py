import sqlite3

import ttkbootstrap as tb
from tkinter import filedialog, messagebox
from tkinter import *
import os
import hashlib
import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from datetime import datetime


from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

import alert

conn = sqlite3.connect('user.db')
c = conn.cursor()

# Create a table if it doesn't already exist
c.execute('''CREATE TABLE IF NOT EXISTS emails
             (email text)''')

root = tb.Window(themename="superhero")
root.title("Integrity Checker")

mydir =''

wb = Workbook()
ws = wb.active
row_num = 2

# set the column headers
ws['A1'] = 'File Path'
ws['B1'] = 'Event'
ws['C1'] = 'Time'


def calculate_file_hash(directory, filepath):
    try:
        with open(os.path.join(directory, filepath), 'rb') as fh:
            hasher = hashlib.sha512()
            hasher.update(fh.read())
            hash_value = hasher.hexdigest()
        return hash_value
    except PermissionError:
            print(f"Error: Permission denied for file {filepath}")
            return None


def addPath():
    global mydir
    mydir = filedialog.askdirectory()
    #with open('dir_selection.txt', 'w') as f:
     #   f.write(mydir)
   # print(mydir)
    files = os.listdir(mydir)

    baseline_file = 'baseline.txt'
    if os.path.exists(baseline_file):
        os.remove(baseline_file)

    with open(baseline_file, 'a') as f:
        for file in files:
            hash_value = calculate_file_hash(mydir, file)
            if hash_value:
                f.write(f"{file}|{hash_value}\n")
            else:
                f.write(f"{file}|ERROR\n")
            #print(hash_value)


class FileHandler(FileSystemEventHandler):
    def on_created(self, event):
        print(f"{event.src_path} has been created!")

def monitor():
    try:
        global row_num
        root.iconify()
        file_hash_dict = {}
        files_to_delete = []
        #files_to_update=[]

        # Load file|hash from baseline.txt and store them in a dictionary
        with open('baseline.txt', 'r') as f:
            file_paths_and_hashes = f.readlines()

        for f in file_paths_and_hashes:
            file_path, file_hash = f.strip().split('|')
            file_hash_dict[file_path] = file_hash

        #mydir = '/path/to/monitor'
        current_directory = os.getcwd()
        excel = os.path.join(current_directory, 'file_changes.xlsx')

        while True:
            observer = Observer()
            observer.schedule(FileHandler(), mydir, recursive=False)
            observer.start()
            #test = email
            x=1
            time.sleep(1)
            files = os.listdir(mydir)
            for file in files:
                hash_value = calculate_file_hash(mydir, file)
                file_path = os.path.join(mydir, file)
                if hash_value not in file_hash_dict.values():
                    print(f"{file_path} file created or edited ")
                    ws[f'A{row_num}'] = file_path
                    ws[f'B{row_num}'] = 'Created or Edited'
                    ws[f'C{row_num}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    row_num += 1
                    wb.save(excel)
                    file_hash_dict[file] = hash_value
                    #alert.new_send(test,file_path)
                   # print(file_hash_dict)
                    x=0

            files = os.listdir(mydir)
            for key in file_hash_dict.keys():
                if key not in files:
                    #print(f"{key} has been deleted!")
                    #rem=key
                    ws[f'A{row_num}'] = file_path
                    ws[f'B{row_num}'] = 'Deleted'
                    ws[f'C{row_num}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    row_num += 1
                    wb.save(excel)
                    #alert.send_email(test,rem)
                    files_to_delete.append(key)
                    x=0

            for key in files_to_delete:
                del file_hash_dict[key]
                #print(f"file_hash_dict: {file_hash_dict}")

            if x==0:
                with open('baseline.txt', 'w') as f:
                    for file in files:
                        hash_value = calculate_file_hash(mydir, file)
                        if hash_value:
                            f.write(f"{file}|{hash_value}\n")
                        else:
                            f.write(f"{file}|ERROR\n")
            observer.stop()
            monitor()

    except Exception as e:
        print(e)
        messagebox.showerror("Error", "Select a folder then begin monitoring ")
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


def firstScreen():
    root.geometry("300x300")

    lbl = Label(root, text="Enter email ")
    lbl.config(anchor=CENTER)
    lbl.pack()

    txt = Entry(root, width=20)
    txt.pack()
    txt.focus()

    def insert():
        global email
        email = txt.get()
        if not email:
            messagebox.showerror("Error", "Missing email ")
            return
            # Insert the email into the database
        #c.execute("INSERT INTO emails (email) VALUES (?)", (email,))
        c.execute("DELETE FROM emails")
        conn.commit()
        c.execute("INSERT OR IGNORE INTO emails (email) VALUES (?)", (email,))
        conn.commit()
        conn.close()
        monitor()


    btn = tb.Button(root,bootstyle="success", text="Add new Path", width=20, command=addPath)
    btn.pack(pady=20)

    btn = tb.Button(root,bootstyle="success", text="Begin Monitoring", width=20, command=insert)
    btn.pack(pady=10)


   # btn = tb.Button(root,bootstyle="Danger", text="Stop Monitoring", width=20, command=stop)
    #btn.pack(pady=10)

#import getpass
#print(getpass.getuser())
firstScreen()
class FileHandler(FileSystemEventHandler):
    def on_created(self, event):
        message = f"{event.src_path} has been created!"
        print(message)
      #  cursor = conn.cursor()
        testt = email
        #alert.new_send(testt, message)


root.mainloop()